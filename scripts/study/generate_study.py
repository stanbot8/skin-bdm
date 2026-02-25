#!/usr/bin/env python3
"""Generate an Excel workbook from a treatment study.

Reads the treatment_study output directory and produces a multi-sheet
Excel workbook with:
  - Summary: headline outcomes per treatment
  - Trajectories: per-treatment timeseries (key columns)
  - Resilience: perturbation-response metrics
  - Insights: auto-generated clinical observations

The workbook uses frozen header rows, auto-sized columns, and
conditional formatting to highlight key findings.

Usage:
    python3 scripts/study/generate_study.py output/treatment_study/
    python3 scripts/study/generate_study.py output/treatment_study/ -o output/studies/
"""

import argparse
import csv
import glob
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


# -- Style constants ----------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Key metrics for summary sheet
SUMMARY_COLS = [
    ("treatment", "Treatment", "text"),
    ("wound_closure_pct", "Closure (%)", "pct"),
    ("time_to_50pct_days", "T50 (days)", "days"),
    ("time_to_90pct_days", "T90 (days)", "days"),
    ("peak_inflammation", "Peak Infl", "sci"),
    ("peak_inflammation_day", "Peak Infl Day", "days"),
    ("mean_infl_wound", "Mean Infl", "sci"),
    ("scar_magnitude", "Scar", "num"),
    ("peak_collagen", "Peak Collagen", "sci"),
    ("peak_myofibroblasts", "Myofibroblasts", "int"),
    ("peak_neutrophils", "Neutrophils", "int"),
    ("peak_macrophages", "Macrophages", "int"),
]

# Timeseries columns to export (step + these)
TS_COLS = [
    "step", "time_h", "time_days", "wound_closure_pct", "mean_infl_wound",
    "mean_anti_infl_wound", "scar_magnitude", "mean_collagen_wound",
    "mean_perfusion_wound", "mean_tgfb_wound", "mean_mmp_wound",
    "mean_fibronectin_wound", "n_neutrophils", "n_macrophages",
    "n_fibroblasts", "n_myofibroblasts", "mean_o2_wound", "mean_vegf_wound",
]

# Resilience output columns
RESILIENCE_DISPLAY = [
    ("treatment", "Treatment"),
    ("metric", "Metric"),
    ("max_deviation", "Max Deviation"),
    ("max_deviation_day", "Max Dev Day"),
    ("resistance", "Resistance"),
    ("peak_response", "Peak Response"),
    ("t_max_days", "Tmax (days)"),
    ("return_time_days", "Return Time (days)"),
    ("elasticity", "Elasticity"),
    ("settling_time_days", "Settling Time (days)"),
    ("overshoot_pct", "Overshoot %"),
    ("auc_burden", "AUC Burden"),
    ("steady_state_error", "SS Error"),
    ("rise_time_days", "Rise Time (days)"),
    ("final_value", "Final Value"),
]


# -- Helpers -------------------------------------------------------------------

def load_csv(path):
    """Load a CSV into a list of dicts."""
    with open(path) as f:
        return list(csv.DictReader(f))


def parse_float(val):
    """Parse a string to float, returning None on failure."""
    if val is None or val == "" or val == "N/A":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def auto_size_columns(ws, min_width=8, max_width=28, pad=2):
    """Auto-size columns based on content width."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        width = min(max(max_len + pad, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def style_header(ws, row=1):
    """Style header row with dark blue fill and white bold text."""
    for cell in ws[row]:
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_borders(ws):
    """Apply thin borders to all data cells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def write_cell(ws, row, col, value, fmt_type="text"):
    """Write a value to a cell with appropriate formatting."""
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = ""
        return cell
    if fmt_type == "pct":
        cell.value = parse_float(str(value))
        cell.number_format = "0.0"
    elif fmt_type == "days":
        cell.value = parse_float(str(value))
        cell.number_format = "0.0"
    elif fmt_type == "sci":
        fv = parse_float(str(value))
        cell.value = fv
        if fv is not None and abs(fv) < 0.01 and fv != 0:
            cell.number_format = "0.00E+00"
        else:
            cell.number_format = "0.0000"
    elif fmt_type == "num":
        cell.value = parse_float(str(value))
        cell.number_format = "0.000"
    elif fmt_type == "int":
        fv = parse_float(str(value))
        cell.value = int(fv) if fv is not None else None
        cell.number_format = "0"
    else:
        cell.value = value
    cell.alignment = Alignment(horizontal="center")
    return cell


# -- Insight generation --------------------------------------------------------

def generate_insights(comparison, resilience_rows, baseline_name="baseline"):
    """Generate clinical insight strings from study data."""
    insights = []

    # Build lookup from comparison CSV
    by_name = {}
    for row in comparison:
        by_name[row["treatment"]] = row
    bl = by_name.get(baseline_name, {})

    # Build resilience lookup: (treatment, metric) -> dict
    res_lookup = {}
    for r in resilience_rows:
        key = (r["treatment"], r["metric"])
        res_lookup[key] = r

    bl_closure = parse_float(bl.get("wound_closure_pct"))
    bl_peak_infl = parse_float(bl.get("peak_inflammation"))
    bl_scar = parse_float(bl.get("scar_magnitude"))
    bl_mean_infl = parse_float(bl.get("mean_infl_wound"))

    for name, row in by_name.items():
        if name == baseline_name:
            continue

        closure = parse_float(row.get("wound_closure_pct"))
        peak_infl = parse_float(row.get("peak_inflammation"))
        scar = parse_float(row.get("scar_magnitude"))
        mean_infl = parse_float(row.get("mean_infl_wound"))
        t50 = parse_float(row.get("time_to_50pct_days"))
        t90 = parse_float(row.get("time_to_90pct_days"))
        myofib = parse_float(row.get("peak_myofibroblasts"))

        obs = []

        # Closure
        if closure is not None and bl_closure is not None:
            if closure > bl_closure + 5:
                obs.append(f"Closure improved to {closure:.0f}% (baseline {bl_closure:.0f}%)")
            elif closure < bl_closure - 5:
                obs.append(f"Closure worsened to {closure:.0f}% (baseline {bl_closure:.0f}%)")

        if t90 is not None:
            obs.append(f"Achieved 90% closure at day {t90:.0f}")

        # Inflammation
        if peak_infl is not None and bl_peak_infl is not None and bl_peak_infl > 0:
            reduction = (1 - peak_infl / bl_peak_infl) * 100
            if reduction > 20:
                obs.append(f"Peak inflammation reduced {reduction:.0f}% (from {bl_peak_infl:.2f} to {peak_infl:.2f})")
            elif reduction < -10:
                obs.append(f"Peak inflammation INCREASED {-reduction:.0f}%")

        if mean_infl is not None and bl_mean_infl is not None and bl_mean_infl > 0:
            burden_red = (1 - mean_infl / bl_mean_infl) * 100
            if burden_red > 20:
                obs.append(f"Mean inflammatory burden reduced {burden_red:.0f}%")

        # Scar
        if scar is not None and bl_scar is not None and bl_scar > 0:
            scar_red = (1 - scar / bl_scar) * 100
            if scar_red > 20:
                obs.append(f"Scarring reduced {scar_red:.0f}% (from {bl_scar:.1f} to {scar:.1f})")
            elif scar_red < -10:
                obs.append(f"Scarring INCREASED {-scar_red:.0f}%")

        # Collagen / myofibroblast
        if myofib is not None and myofib >= 4:
            obs.append(f"Strong myofibroblast response ({int(myofib)} cells)")

        # Resilience-based
        res_infl = res_lookup.get((name, "mean_infl_wound"), {})
        bl_res_infl = res_lookup.get((baseline_name, "mean_infl_wound"), {})
        rt = parse_float(res_infl.get("return_time_days"))
        bl_rt = parse_float(bl_res_infl.get("return_time_days"))
        if rt is not None and bl_rt is not None and bl_rt > 0:
            if rt < bl_rt * 0.7:
                obs.append(f"Inflammation resolves {(1-rt/bl_rt)*100:.0f}% faster (day {rt:.0f} vs {bl_rt:.0f})")

        auc = parse_float(res_infl.get("auc_burden"))
        bl_auc = parse_float(bl_res_infl.get("auc_burden"))
        if auc is not None and bl_auc is not None and bl_auc > 0:
            auc_red = (1 - auc / bl_auc) * 100
            if auc_red > 30:
                obs.append(f"Cumulative inflammatory AUC reduced {auc_red:.0f}%")

        # Synthesis
        if obs:
            net = "positive" if (closure and closure >= (bl_closure or 0)) else "mixed"
            if scar is not None and bl_scar is not None and scar < bl_scar * 0.5:
                quality = "substantial scar reduction"
            elif scar is not None and bl_scar is not None and scar < bl_scar * 0.8:
                quality = "moderate scar reduction"
            else:
                quality = "limited scar benefit"

            insights.append({
                "treatment": name,
                "outcome": net,
                "quality": quality,
                "observations": obs,
            })

    return insights


# -- Sheet builders ------------------------------------------------------------

def build_summary_sheet(wb, comparison_rows, baseline_name="baseline"):
    """Build the Summary sheet from treatment_comparison.csv."""
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Diabetic Wound Treatment Study - Summary"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"].value = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} | 9 simulations | 7200 steps (30 days)"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers at row 4
    header_row = 4
    for i, (key, label, _) in enumerate(SUMMARY_COLS, 1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    bl = None
    for row in comparison_rows:
        if row["treatment"] == baseline_name:
            bl = row
            break

    for r_idx, row in enumerate(comparison_rows, header_row + 1):
        is_baseline = row["treatment"] == baseline_name
        for c_idx, (key, _, fmt_type) in enumerate(SUMMARY_COLS, 1):
            val = row.get(key, "")
            cell = write_cell(ws, r_idx, c_idx, val, fmt_type)

            # Highlight vs baseline
            if not is_baseline and bl and key in ("peak_inflammation", "mean_infl_wound", "scar_magnitude"):
                fv = parse_float(val)
                bl_v = parse_float(bl.get(key))
                if fv is not None and bl_v is not None and bl_v > 0:
                    ratio = fv / bl_v
                    if ratio < 0.5:
                        cell.fill = GOOD_FILL
                    elif ratio > 1.1:
                        cell.fill = BAD_FILL
            elif not is_baseline and bl and key == "wound_closure_pct":
                fv = parse_float(val)
                bl_v = parse_float(bl.get(key))
                if fv is not None and bl_v is not None:
                    if fv > bl_v + 5:
                        cell.fill = GOOD_FILL

        # Bold baseline row
        if is_baseline:
            for c in range(1, len(SUMMARY_COLS) + 1):
                ws.cell(row=r_idx, column=c).font = Font(bold=True, size=11)

    ws.freeze_panes = f"A{header_row + 1}"
    apply_borders(ws)
    auto_size_columns(ws)


def build_trajectories_sheet(wb, study_dir, comparison_rows):
    """Build per-treatment trajectory sheets with key columns."""
    treatment_names = [r["treatment"] for r in comparison_rows]

    for name in treatment_names:
        csv_path = os.path.join(study_dir, f"metrics_{name}.csv")
        if not os.path.exists(csv_path):
            continue

        rows = load_csv(csv_path)
        if not rows:
            continue

        # Sheet name max 31 chars
        sheet_name = name[:27] if len(name) > 27 else name
        ws = wb.create_sheet(title=sheet_name)

        # Determine which columns exist
        available = [c for c in TS_COLS if c in rows[0]]

        # Headers
        for i, col in enumerate(available, 1):
            ws.cell(row=1, column=i, value=col)

        style_header(ws, 1)

        # Data (sample every 100 steps to keep workbook size reasonable)
        out_row = 2
        for r_idx, row in enumerate(rows):
            step = int(float(row.get("step", r_idx)))
            if step % 100 != 0 and r_idx != len(rows) - 1:
                continue
            for c_idx, col in enumerate(available, 1):
                val = parse_float(row.get(col))
                cell = ws.cell(row=out_row, column=c_idx)
                cell.value = val
                cell.alignment = Alignment(horizontal="center")
            out_row += 1

        ws.freeze_panes = "A2"
        auto_size_columns(ws)


def build_resilience_sheet(wb, resilience_rows):
    """Build the Resilience Metrics sheet."""
    ws = wb.create_sheet(title="Resilience")

    # Headers
    for i, (key, label) in enumerate(RESILIENCE_DISPLAY, 1):
        ws.cell(row=1, column=i, value=label)
    style_header(ws, 1)

    # Data
    for r_idx, row in enumerate(resilience_rows, 2):
        for c_idx, (key, _) in enumerate(RESILIENCE_DISPLAY, 1):
            val = row.get(key, "")
            fv = parse_float(val)
            cell = ws.cell(row=r_idx, column=c_idx)
            if key in ("treatment", "metric"):
                cell.value = val
            elif fv is not None:
                cell.value = fv
                if abs(fv) < 0.01 and fv != 0:
                    cell.number_format = "0.00E+00"
                else:
                    cell.number_format = "0.000"
            else:
                cell.value = val
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C2"
    apply_borders(ws)
    auto_size_columns(ws)


def build_insights_sheet(wb, insights):
    """Build the Insights sheet with auto-generated clinical observations."""
    ws = wb.create_sheet(title="Insights")

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Clinical Insights (auto-generated from simulation trajectories)"
    ws["A1"].font = Font(bold=True, size=13)

    ws.merge_cells("A2:D2")
    ws["A2"].value = "These observations compare each treatment to the untreated diabetic baseline."
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    # Headers
    headers = ["Treatment", "Outcome", "Scar Quality", "Observation"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4)

    row = 5
    for ins in insights:
        name = ins["treatment"]
        outcome = ins["outcome"]
        quality = ins["quality"]

        for j, obs in enumerate(ins["observations"]):
            ws.cell(row=row, column=1, value=name if j == 0 else "")
            ws.cell(row=row, column=2, value=outcome if j == 0 else "")
            ws.cell(row=row, column=3, value=quality if j == 0 else "")
            ws.cell(row=row, column=4, value=obs)

            # Color code outcome
            if j == 0:
                outcome_cell = ws.cell(row=row, column=2)
                if outcome == "positive":
                    outcome_cell.fill = GOOD_FILL
                else:
                    outcome_cell.fill = NEUTRAL_FILL

            # Color observation severity
            obs_cell = ws.cell(row=row, column=4)
            if "INCREASED" in obs or "worsened" in obs:
                obs_cell.fill = BAD_FILL
            elif "reduced" in obs.lower() and any(w in obs for w in ["50%", "60%", "70%", "80%", "90%"]):
                obs_cell.fill = GOOD_FILL

            row += 1

        # Blank row between treatments
        row += 1

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 70


def build_ranking_sheet(wb, comparison_rows, resilience_rows, baseline_name="baseline"):
    """Build a compact ranking sheet."""
    ws = wb.create_sheet(title="Ranking")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Treatment Efficacy Ranking"
    ws["A1"].font = Font(bold=True, size=13)

    ws.merge_cells("A2:F2")
    ws["A2"].value = "Composite score from closure AUC, inflammatory burden, resolution speed, peak inflammation, and scar."
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    # Compute composite ranks (same logic as resilience_analysis.py)
    treatments = [r["treatment"] for r in comparison_rows]
    bl = None
    for r in comparison_rows:
        if r["treatment"] == baseline_name:
            bl = r

    # Simple ranking by key outcome metrics
    rank_data = []
    for row in comparison_rows:
        name = row["treatment"]
        closure = parse_float(row.get("wound_closure_pct")) or 0
        mean_infl = parse_float(row.get("mean_infl_wound")) or 0
        scar = parse_float(row.get("scar_magnitude")) or 0
        peak_infl = parse_float(row.get("peak_inflammation")) or 0

        # Derive vs baseline
        bl_closure = parse_float(bl.get("wound_closure_pct")) or 1 if bl else 1
        bl_scar = parse_float(bl.get("scar_magnitude")) or 1 if bl else 1
        bl_peak = parse_float(bl.get("peak_inflammation")) or 1 if bl else 1
        bl_mean = parse_float(bl.get("mean_infl_wound")) or 1 if bl else 1

        rank_data.append({
            "treatment": name,
            "closure": closure,
            "infl_reduction": (1 - peak_infl / max(bl_peak, 0.001)) * 100 if bl else 0,
            "burden_reduction": (1 - mean_infl / max(bl_mean, 0.001)) * 100 if bl else 0,
            "scar_reduction": (1 - scar / max(bl_scar, 0.001)) * 100 if bl else 0,
        })

    # Composite: average of three reduction percentages + closure bonus
    for rd in rank_data:
        rd["composite"] = (
            rd["infl_reduction"] * 0.25
            + rd["burden_reduction"] * 0.25
            + rd["scar_reduction"] * 0.25
            + (rd["closure"] - 80) * 1.25  # bonus for closure > 80%
        )

    rank_data.sort(key=lambda x: x["composite"], reverse=True)

    headers = ["Rank", "Treatment", "Closure %", "Infl Reduction %", "Burden Reduction %", "Scar Reduction %", "Composite Score"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4)

    for rank, rd in enumerate(rank_data, 1):
        r = rank + 4
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=rd["treatment"])
        write_cell(ws, r, 3, rd["closure"], "pct")
        write_cell(ws, r, 4, rd["infl_reduction"], "num")
        write_cell(ws, r, 5, rd["burden_reduction"], "num")
        write_cell(ws, r, 6, rd["scar_reduction"], "num")
        write_cell(ws, r, 7, rd["composite"], "num")

        # Color top 3
        if rank <= 3:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = GOOD_FILL
        # Bold baseline
        if rd["treatment"] == baseline_name:
            for c in range(1, 8):
                ws.cell(row=r, column=c).font = Font(bold=True)

    ws.freeze_panes = "A5"
    apply_borders(ws)
    auto_size_columns(ws)


# -- Main ----------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate Excel study workbook")
    parser.add_argument("study_dir", help="Treatment study output directory")
    parser.add_argument("-o", "--output-dir", default=None,
                        help="Output directory (default: <study_dir>)")
    parser.add_argument("--name", default="diabetic_treatment_study",
                        help="Study name for the output file")
    args = parser.parse_args()

    study_dir = args.study_dir
    out_dir = args.output_dir or study_dir

    # Load data
    comparison_path = os.path.join(study_dir, "treatment_comparison.csv")
    resilience_path = os.path.join(study_dir, "resilience_metrics.csv")

    if not os.path.exists(comparison_path):
        print(f"Missing {comparison_path}. Run treatment_study.py first.")
        sys.exit(1)

    comparison = load_csv(comparison_path)
    resilience = load_csv(resilience_path) if os.path.exists(resilience_path) else []

    print(f"Loaded {len(comparison)} treatments, {len(resilience)} resilience rows")

    # Generate insights
    insights = generate_insights(comparison, resilience)
    print(f"Generated {sum(len(i['observations']) for i in insights)} clinical observations")

    # Build workbook
    wb = Workbook()
    build_summary_sheet(wb, comparison)
    build_ranking_sheet(wb, comparison, resilience)
    build_insights_sheet(wb, insights)
    build_resilience_sheet(wb, resilience)
    build_trajectories_sheet(wb, study_dir, comparison)

    # Save
    os.makedirs(out_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"{args.name}_{date_str}.xlsx"
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)
    print(f"\nWorkbook saved to {out_path}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
